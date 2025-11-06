from openpyxl import Workbook, load_workbook
import datetime
import json
import os
import re
from fpdf import FPDF
from pathlib import Path
import datetime

import pandas as pd




def excel___(response_text: str, output_file=str(Path.cwd())+ "/files/plan" + str(datetime.datetime.now()) + ".xlsx"):
    try:
        # Nettoyage du texte JSON (enlève les ```json ... ``` éventuels)
        cleaned = re.sub(r"^```[a-zA-Z]*\n?", "", response_text.strip())
        cleaned = re.sub(r"```$", "", cleaned.strip())
        print(cleaned)
        # Conversion en dict Python
        response_json = json.loads(cleaned)

        # Création du fichier Excel
        with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
            for key, value in response_json.items():
                sheet_name = key[:31]  # max 31 caractères pour Excel
                if isinstance(value, list):
                    pd.DataFrame(value).to_excel(writer, sheet_name=sheet_name, index=False)
                elif isinstance(value, dict):
                    pd.DataFrame([value]).to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    pd.DataFrame([{"valeur": value}]).to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Fichier Excel généré : {output_file}")
        return output_file

    except Exception as e:
        print(f"Erreur lors de la création Excel : {e}")
        print("Tentative de génération d’un PDF à la place...")

        # Si Excel échoue → générer un PDF
        pdf_file = str(Path(output_file).with_suffix(".pdf"))
        try:
            generate_pdf_fallback(cleaned, pdf_file)
            print(f"Fichier PDF de secours généré : {pdf_file}")
            return pdf_file
        except Exception as e2:
            print(f"Échec aussi de la génération PDF : {e2}")
            raise e2


def excel(response_text: str, output_file=str(Path.cwd())+ "/files/plan" + str(datetime.datetime.now()) + ".xlsx"):
    try:
        # --- Nettoyage du texte JSON ---
        cleaned = re.sub(r"^```[a-zA-Z]*\n?", "", response_text.strip())
        cleaned = re.sub(r"```$", "", cleaned.strip())

        # --- Conversion en dictionnaire Python ---
        response_json = json.loads(cleaned)

        # --- Création du fichier Excel ---
        with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
            workbook = writer.book
            wrap_format = workbook.add_format({"text_wrap": True, "valign": "top"})
            
            for key, value in response_json.items():
                sheet_name = str(key)[:31]  # nom de feuille Excel max 31 caractères
                
                # Cas 1 : liste (soit de chaînes, soit de dictionnaires)
                if isinstance(value, list):
                    if all(isinstance(v, dict) for v in value):
                        # Liste de dictionnaires → une ligne par dictionnaire
                        formatted_list = []
                        for item in value:
                            if isinstance(item, dict):
                                lines = [f"{k}: {v}" for k, v in item.items()]
                                formatted_list.append("\n".join(lines))
                            else:
                                formatted_list.append(str(item))
                        df = pd.DataFrame({"valeurs": formatted_list})

                    elif all(isinstance(v, str) for v in value):
                        # Liste de chaînes → une cellule avec retour à la ligne
                        joined_text = "\n".join(value)
                        df = pd.DataFrame({"valeurs": [joined_text]})

                    else:
                        # Liste mixte → convertir proprement
                        df = pd.DataFrame({"valeurs": [str(v) for v in value]})

                # Cas 2 : dictionnaire → clé | valeur
                elif isinstance(value, dict):
                    df = pd.DataFrame(list(value.items()), columns=["clé", "valeur"])

                # Cas 3 : valeur simple
                else:
                    df = pd.DataFrame({"valeurs": [value]})

                # Écriture dans Excel
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]

                # Largeur automatique + format avec retour à la ligne
                for col_num, col in enumerate(df.columns):
                    max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.set_column(col_num, col_num, max_len, wrap_format)

        print(f"Fichier Excel généré : {output_file}")
        return output_file
    except Exception as e:
        print(f"Erreur lors de la création Excel : {e}")
        print("Tentative de génération d’un PDF à la place...")

        # Si Excel échoue → générer un PDF
        pdf_file = str(Path(output_file).with_suffix(".pdf"))
        try:
            generate_pdf_fallback(cleaned, pdf_file)
            print(f"Fichier PDF de secours généré : {pdf_file}")
            return pdf_file
        except Exception as e2:
            print(f"Échec aussi de la génération PDF : {e2}")
            raise e2


def generate_pdf_fallback(json_text: str, pdf_path: str):
    """Crée un PDF contenant le JSON brut (lisible, indenté)."""
    data = json.loads(json_text)
    pretty_json = json.dumps(data, indent=2, ensure_ascii=False)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Courier", size=10)

    # Découper le JSON en lignes pour éviter le dépassement
    for line in pretty_json.split("\n"):
        pdf.multi_cell(0, 6, line)

    pdf.output(pdf_path)

def create_excel_plan(ai_response: str):
    """
    Crée un fichier Excel contenant le plan financier généré par l’IA.
    """
    if not os.path.exists("files"):
        os.makedirs("files")

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan Financier"

    #ws["A1"] = "Planification Financière Personnalisée"
    #ws["A3"] = "Recommandations IA"
    ws["A1"] = ai_response

    filename = "plan_financier.xlsx"


    filepath = os.path.join("files", filename)
    wb.save(filepath)
    return filename
